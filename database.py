import aiosqlite
import csv
import io
import logging
from pathlib import Path
from datetime import datetime
from config import DB_NAME, ADMIN_IDS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

ENV_SUPER_ADMIN_IDS = ADMIN_IDS if ADMIN_IDS else []

# Таймаут ожидания снятия блокировки БД другим соединением (в секундах).
# Эквивалентен PRAGMA busy_timeout и уменьшает вероятность "database is locked"
# при конкурентной записи нескольких обработчиков одновременно.
DB_BUSY_TIMEOUT_SECONDS = 5


def get_connection() -> aiosqlite.Connection:
    """Единая точка открытия соединения с БД с настроенным busy_timeout."""
    return aiosqlite.connect(DB_NAME, timeout=DB_BUSY_TIMEOUT_SECONDS)


def _quote_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'

async def _sync_archive_schema(db: aiosqlite.Connection):
    claims_info = await (await db.execute("PRAGMA table_info(claims)")).fetchall()
    archive_info = await (await db.execute("PRAGMA table_info(claims_archive)")).fetchall()

    claims_columns = {row[1]: row for row in claims_info}
    archive_columns = {row[1] for row in archive_info}
    missing_columns = [name for name in claims_columns if name not in archive_columns]

    for column_name in missing_columns:
        _, _, column_type, not_null, default_value, _ = claims_columns[column_name]
        quoted_name = _quote_identifier(column_name)

        parts = [f"ALTER TABLE claims_archive ADD COLUMN {quoted_name}"]
        if column_type:
            parts.append(column_type)
        # SQLite не позволяет безопасно добавить NOT NULL колонку без DEFAULT
        # в уже существующую таблицу с данными.
        if not_null and default_value is not None:
            parts.append("NOT NULL")
        if default_value is not None:
            parts.append(f"DEFAULT {default_value}")

        await db.execute(" ".join(parts))

async def archive_old_claims(days: int = 365):
    async with get_connection() as db:
        await _sync_archive_schema(db)
        await db.execute(f"""
            INSERT INTO claims_archive (
                id, display_id, user_id, category, sub_category, brand,
                defect_desc, purchase_date, client_wish, photo_id, status,
                admin_comment, admin_name, client_name, tg_name, created_at,
                payment_method, competitor_offer, chat_locked, buyout_amount
            )
            SELECT
                id, display_id, user_id, category, sub_category, brand,
                defect_desc, purchase_date, client_wish, photo_id, status,
                admin_comment, admin_name, client_name, tg_name, created_at,
                payment_method, competitor_offer, chat_locked, buyout_amount
            FROM claims
            WHERE date(created_at) < date('now', '-{days} days')
        """)
        await db.execute(f"""
            DELETE FROM claims 
            WHERE date(created_at) < date('now', '-{days} days')
        """)
        await db.commit()
        cursor = await db.execute("SELECT changes()")
        archived_count = (await cursor.fetchone())[0]
        return archived_count

async def init_db():
    async with get_connection() as db:
        # WAL позволяет читателям не блокироваться на время записи и наоборот,
        # что вместе с busy_timeout (см. get_connection) заметно снижает риск
        # "database is locked" при конкурентных обращениях к SQLite.
        await db.execute("PRAGMA journal_mode=WAL")
        await db.execute("PRAGMA busy_timeout=5000")
        await db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                role TEXT DEFAULT 'user'
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                display_id TEXT UNIQUE,
                user_id INTEGER,
                category TEXT,
                sub_category TEXT,
                brand TEXT,
                defect_desc TEXT,
                purchase_date TEXT,
                client_wish TEXT,
                photo_id TEXT,
                status TEXT DEFAULT 'pending',
                admin_comment TEXT,
                admin_name TEXT,
                client_name TEXT DEFAULT 'Не указано',
                tg_name TEXT DEFAULT '',
                payment_method TEXT,
                competitor_offer TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_id INTEGER,
                action TEXT,
                target_id INTEGER,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS updates_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                update_type TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS claim_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                claim_id INTEGER,
                display_id TEXT,
                old_status TEXT,
                new_status TEXT,
                admin_id INTEGER,
                admin_name TEXT,
                comment TEXT,
                changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS claims_archive AS SELECT * FROM claims WHERE 1=0
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS claim_counters (
                category TEXT PRIMARY KEY,
                last_number INTEGER DEFAULT 0
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version TEXT PRIMARY KEY,
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("CREATE INDEX IF NOT EXISTS idx_claims_display_id ON claims(display_id)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_claims_user_id ON claims(user_id)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_claims_category_status ON claims(category, status)")
        await db.execute("""
            INSERT OR IGNORE INTO claim_counters (category, last_number) VALUES ('tech', 0)
        """)
        await db.execute("""
            INSERT OR IGNORE INTO claim_counters (category, last_number) VALUES ('acc', 0)
        """)
        await db.execute("""
            INSERT OR IGNORE INTO claim_counters (category, last_number) VALUES ('tradein', 0)
        """)
        await db.execute("""
            INSERT OR IGNORE INTO claim_counters (category, last_number) VALUES ('complaint', 0)
        """)
        if ENV_SUPER_ADMIN_IDS:
            for admin_id in ENV_SUPER_ADMIN_IDS:
                await db.execute(
                    "INSERT OR REPLACE INTO users (user_id, role) VALUES (?, ?)",
                    (admin_id, 'super_admin')
                )
        await db.commit()
    logger.info("Database schema initialized (WAL mode, busy_timeout=%ss)", DB_BUSY_TIMEOUT_SECONDS)
    await apply_migrations()
    try:
        from utils.bot_config import ensure_bot_config_seeded
        from utils.tz import refresh_display_tz_from_settings
        await ensure_bot_config_seeded()
        await refresh_display_tz_from_settings()
    except Exception as exc:
        logger.error("Failed to seed bot config / refresh timezone: %s", exc)

async def apply_migrations():
    migrations_dir = Path(__file__).resolve().parent / "migrations"
    if not migrations_dir.exists():
        return

    migration_files = sorted(migrations_dir.glob("*.sql"))
    if not migration_files:
        return

    applied_count = 0
    async with get_connection() as db:
        for migration_file in migration_files:
            version = migration_file.name
            cursor = await db.execute(
                "SELECT 1 FROM schema_migrations WHERE version = ?",
                (version,)
            )
            exists = await cursor.fetchone()
            if exists:
                continue

            sql = migration_file.read_text(encoding="utf-8").strip()
            if sql:
                try:
                    await db.executescript(sql)
                except aiosqlite.OperationalError as exc:
                    # Позволяем идемпотентно пережить миграции вида ALTER TABLE ... ADD COLUMN
                    # если колонка уже создана в новой схеме.
                    if "duplicate column name" not in str(exc).lower():
                        raise
                    logger.info("Migration %s already applied to schema (columns exist), marking as applied", version)
            await db.execute(
                "INSERT INTO schema_migrations (version) VALUES (?)",
                (version,)
            )
            applied_count += 1
            logger.info("Applied migration: %s", version)
        await db.commit()
    if applied_count == 0:
        logger.info("No new migrations to apply, schema is up to date")
    else:
        logger.info("Applied %s new migration(s)", applied_count)

async def get_user_role(user_id: int) -> str:
    async with get_connection() as db:
        cursor = await db.execute("SELECT role FROM users WHERE user_id = ?", (user_id,))
        res = await cursor.fetchone()
        return res[0] if res else 'user'

async def set_user_role(user_id: int, role: str):
    async with get_connection() as db:
        await db.execute(
            "INSERT OR REPLACE INTO users (user_id, role) VALUES (?, ?)",
            (user_id, role)
        )
        await db.commit()
    logger.info("User role updated: user_id=%s new_role=%s", user_id, role)

async def log_action(admin_id: int, action: str, target_id: int = None):
    async with get_connection() as db:
        await db.execute(
            "INSERT INTO logs (admin_id, action, target_id) VALUES (?, ?, ?)",
            (admin_id, action, target_id)
        )
        await db.commit()

async def log_update(user_id: int, update_type: str):
    async with get_connection() as db:
        await db.execute(
            "INSERT INTO updates_log (user_id, update_type) VALUES (?, ?)",
            (user_id, update_type)
        )
        await db.commit()

async def get_next_display_id(category: str) -> str:
    async with get_connection() as db:
        await db.execute("BEGIN IMMEDIATE")
        cursor = await db.execute(
            "SELECT last_number FROM claim_counters WHERE category = ?",
            (category,)
        )
        row = await cursor.fetchone()
        current = row[0] if row else 0
        next_num = current + 1
        await db.execute(
            "UPDATE claim_counters SET last_number = ? WHERE category = ?",
            (next_num, category)
        )
        await db.commit()
        
        prefix_map = {
            'tech': 'Т',
            'acc': 'А',
            'tradein': 'В',
            'complaint': 'С'
        }
        prefix = prefix_map.get(category, '?')
        
        return f"{prefix}{next_num}"


async def create_claim(data: dict, user_id: int) -> tuple:
    category = data['category']
    display_id = await get_next_display_id(category)
    async with get_connection() as db:
        cursor = await db.execute("""
            INSERT INTO claims (
                display_id, user_id, category, sub_category, brand, 
                defect_desc, purchase_date, client_wish, photo_id, client_name, tg_name,
                payment_method, competitor_offer
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            display_id, user_id, data['category'], data.get('sub_category'), 
            data.get('brand'), data.get('defect'), data.get('purchase_date'), 
            data.get('client_wish'), data['photo'], data.get('client_name', 'Не указано'),
            data.get('tg_name', ''), data.get('payment_method'), data.get('competitor_offer')
        ))
        claim_id = cursor.lastrowid
        # Системная запись в чат заявки (журнал событий, см. раздел "ЧАТ ЗАЯВКИ" ниже).
        # Вставляется в той же транзакции, что и сама заявка — не требует отдельного
        # соединения и гарантированно появляется одновременно с созданием заявки.
        try:
            await db.execute(
                """INSERT INTO chat_messages (claim_id, sender_id, sender_role, message_type, text)
                   VALUES (?, NULL, 'system', 'system', ?)""",
                (claim_id, "Заявка создана")
            )
        except aiosqlite.OperationalError as exc:
            # Таблица чата появляется миграцией 005 и может отсутствовать только
            # в переходный момент до применения миграций — не должно ломать создание заявки.
            logger.warning("Chat system message on claim create skipped: %s", exc)
        await db.commit()
        logger.info(
            "Claim created: display_id=%s category=%s sub_category=%s user_id=%s",
            display_id, category, data.get('sub_category'), user_id
        )
        return claim_id, display_id

async def get_claim(claim_id: int):
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute("SELECT * FROM claims WHERE id = ?", (claim_id,))
        row = await cursor.fetchone()
        if not row:
            return None
        return dict(row)

async def get_claim_by_display_id(display_id: str):
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute("SELECT * FROM claims WHERE display_id = ?", (display_id,))
        row = await cursor.fetchone()
        if not row:
            return None
        return dict(row)

async def find_claim_by_display_id_or_imei(query: str):
    query_norm = (query or "").strip().upper()
    if not query_norm:
        return None

    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            """
            SELECT *
            FROM claims
            WHERE UPPER(display_id) = ?
               OR (category = 'tech' AND UPPER(brand) LIKE ?)
               OR (category = 'tech' AND UPPER(defect_desc) LIKE ?)
            ORDER BY id DESC
            LIMIT 1
            """,
            (query_norm, f"%IMEI: {query_norm}%", f"%IMEI: {query_norm}%")
        )
        row = await cursor.fetchone()
        if not row:
            return None
        return dict(row)

async def get_claim_by_display_id_for_user(display_id: str, user_id: int):
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            "SELECT * FROM claims WHERE display_id = ? AND user_id = ?",
            (display_id, user_id)
        )
        row = await cursor.fetchone()
        if not row:
            return None
        return dict(row)

async def update_claim_status(claim_id: int, status: str, comment: str = None, admin_name: str = None):
    async with get_connection() as db:
        await db.execute(
            "UPDATE claims SET status = ?, admin_comment = ?, admin_name = ? WHERE id = ?",
            (status, comment, admin_name, claim_id)
        )
        await db.commit()


async def set_claim_buyout_amount(claim_id: int, amount: int) -> bool:
    """Сохраняет фактическую сумму выкупа Trade-in (целое число рублей).

    Возвращает True, если строка заявки обновлена. Не трогает status/admin —
    итог сделки ТТ вторичен относительно админского решения (см. tradein outcome).
    """
    async with get_connection() as db:
        cursor = await db.execute(
            "UPDATE claims SET buyout_amount = ? WHERE id = ?",
            (int(amount), claim_id),
        )
        await db.commit()
        return cursor.rowcount > 0

async def try_update_claim_status(claim_id: int, status: str, comment: str = None, admin_name: str = None) -> tuple:
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        await db.execute("BEGIN IMMEDIATE")
        
        try:
            cursor = await db.execute(
                "SELECT * FROM claims WHERE id = ?",
                (claim_id,)
            )
            row = await cursor.fetchone()
            
            if not row:
                await db.commit()
                return None, None
            
            claim = dict(row)
            current_status = claim.get('status', 'pending')
            
            if current_status != 'pending':
                await db.commit()
                return False, claim
            
            await db.execute(
                "UPDATE claims SET status = ?, admin_comment = ?, admin_name = ? WHERE id = ?",
                (status, comment, admin_name, claim_id)
            )
            await db.commit()
            
            claim['status'] = status
            claim['admin_comment'] = comment
            claim['admin_name'] = admin_name
            
            return True, claim
            
        except Exception:
            await db.rollback()
            raise

async def add_claim_history(claim_id: int, display_id: str, old_status: str, new_status: str, admin_id: int, admin_name: str, comment: str = None):
    async with get_connection() as db:
        await db.execute("""
            INSERT INTO claim_history (claim_id, display_id, old_status, new_status, admin_id, admin_name, comment)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (claim_id, display_id, old_status, new_status, admin_id, admin_name, comment))
        await db.commit()

async def get_claim_history(claim_id: int):
    async with get_connection() as db:
        cursor = await db.execute("""
            SELECT old_status, new_status, admin_name, comment, changed_at 
            FROM claim_history 
            WHERE claim_id = ? 
            ORDER BY changed_at DESC
        """, (claim_id,))
        return await cursor.fetchall()


# ==========================================
# ЧАТ ЗАЯВКИ (ОБСУЖДЕНИЕ)
# ==========================================
# Участники чата заявки НЕ хранятся отдельной таблицей — они вычисляются
# каждый раз из уже существующих данных (claims.user_id, claim_history.admin_id,
# users.role), чтобы не дублировать источник истины и не терять актуальность
# при смене ролей администраторов. Единственное состояние, которое реально
# хранится — сами сообщения (chat_messages) и признак блокировки (claims.chat_locked).

CLAIM_CATEGORY_ADMIN_ROLE = {
    'tech': 'admin_tech',
    'acc': 'admin_acc',
    'tradein': 'admin_tradein',
    'complaint': 'admin_complaint',
}


async def get_claim_responsible_admin_id(claim_id: int):
    """ID администратора, принявшего последнее решение по заявке (по claim_history).

    Пока по заявке нет ни одного решения, единого "ответственного" не существует —
    в этом случае участниками чата со стороны администрации считаются ВСЕ
    администраторы соответствующей роли (см. get_claim_chat_participants).
    """
    async with get_connection() as db:
        cursor = await db.execute(
            """
            SELECT admin_id FROM claim_history
            WHERE claim_id = ? AND admin_id IS NOT NULL
            ORDER BY changed_at DESC, id DESC
            LIMIT 1
            """,
            (claim_id,)
        )
        row = await cursor.fetchone()
        return row[0] if row else None


async def get_claim_responsible_admin_info(claim_id: int):
    """(admin_id, admin_name) администратора, принявшего последнее решение по
    заявке (по claim_history) — используется там, где нужно показать имя
    ответственного сотрудника без похода в Telegram API (например, в
    уведомлении об итоге сделки Trade-in). Возвращает (None, None), если по
    заявке ещё нет ни одного решения."""
    async with get_connection() as db:
        cursor = await db.execute(
            """
            SELECT admin_id, admin_name FROM claim_history
            WHERE claim_id = ? AND admin_id IS NOT NULL
            ORDER BY changed_at DESC, id DESC
            LIMIT 1
            """,
            (claim_id,)
        )
        row = await cursor.fetchone()
        return (row[0], row[1]) if row else (None, None)


async def get_claim_chat_participants(claim: dict) -> dict:
    """Вычисляет участников чата заявки одним компактным набором запросов
    (без N+1: по одному запросу на список админов роли/супер-админов).

    Возвращает {'author_id': int|None, 'admin_ids': set[int], 'super_admin_ids': set[int]}.
    """
    claim_id = claim['id']
    category = claim.get('category')
    role_prefix = CLAIM_CATEGORY_ADMIN_ROLE.get(category)

    responsible_admin_id = await get_claim_responsible_admin_id(claim_id)
    admin_ids = set()
    if responsible_admin_id:
        admin_ids.add(responsible_admin_id)
    elif role_prefix:
        # Заявка ещё не решена — доступ имеют все админы соответствующей роли
        # (именно они и так получают уведомление о новой заявке).
        admin_ids.update(await get_admins_by_role(role_prefix))

    super_admin_ids = set(await get_admins_by_role('super_admin'))

    return {
        'author_id': claim.get('user_id'),
        'admin_ids': admin_ids,
        'super_admin_ids': super_admin_ids,
    }


async def get_claim_chat_role(claim: dict, user_id: int):
    """Роль пользователя в чате конкретной заявки: 'tt' | 'admin' | 'super_admin' | None.

    None означает, что пользователь НЕ является участником чата этой заявки —
    вызывающий код обязан в этом случае вызвать deny_access и не показывать историю.
    Проверка идёт по реальным данным БД, а не по тому, что пользователь передал
    в callback_data (claim_id в колбэке — это только "куда", а не "можно ли").
    """
    if claim.get('user_id') == user_id:
        return 'tt'
    participants = await get_claim_chat_participants(claim)
    if user_id in participants['super_admin_ids']:
        return 'super_admin'
    if user_id in participants['admin_ids']:
        return 'admin'
    return None


async def get_claim_chat_recipient_ids(claim: dict, exclude_id=None) -> set:
    """Полный набор user_id участников чата (для рассылки), одним вычислением."""
    participants = await get_claim_chat_participants(claim)
    recipients = set(participants['admin_ids']) | set(participants['super_admin_ids'])
    if participants['author_id']:
        recipients.add(participants['author_id'])
    recipients.discard(None)
    recipients.discard(exclude_id)
    return recipients


async def add_chat_message(
    claim_id: int,
    sender_id,
    sender_role: str,
    message_type: str,
    text: str = None,
    file_id: str = None,
    reply_to_message_id: int = None,
) -> int:
    async with get_connection() as db:
        cursor = await db.execute(
            """
            INSERT INTO chat_messages (claim_id, sender_id, sender_role, message_type, text, file_id, reply_to_message_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (claim_id, sender_id, sender_role, message_type, text, file_id, reply_to_message_id)
        )
        await db.commit()
        return cursor.lastrowid


async def add_chat_system_message(claim_id: int, text: str) -> int:
    """Автоматическая системная запись о событии жизненного цикла заявки
    (создание, финальное решение) — попадает в общую историю чата заявки."""
    return await add_chat_message(claim_id, sender_id=None, sender_role='system', message_type='system', text=text)


async def get_chat_messages(claim_id: int, limit: int = 200) -> list:
    """История сообщений чата заявки в хронологическом порядке (старые -> новые).

    Один запрос с индексом по (claim_id, created_at) — без N+1. limit защищает
    от неограниченного роста одного текстового сообщения истории у очень старых
    заявок с очень длинной перепиской (см. render_chat_history — дополнительно
    обрезает по количеству символов Telegram).
    """
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            """
            SELECT * FROM chat_messages
            WHERE claim_id = ?
            ORDER BY id ASC
            LIMIT ?
            """,
            (claim_id, limit)
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def get_chat_message(message_id: int):
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute("SELECT * FROM chat_messages WHERE id = ?", (message_id,))
        row = await cursor.fetchone()
        return dict(row) if row else None


async def set_claim_chat_locked(claim_id: int, locked: bool):
    async with get_connection() as db:
        await db.execute(
            "UPDATE claims SET chat_locked = ? WHERE id = ?",
            (1 if locked else 0, claim_id)
        )
        await db.commit()


def is_claim_chat_locked(claim: dict) -> bool:
    return bool(claim.get('chat_locked'))


async def get_admins_by_role(role_prefix: str):
    async with get_connection() as db:
        if role_prefix == 'super_admin':
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'super_admin' ORDER BY user_id")
            rows = await cursor.fetchall()
            db_admins = [row[0] for row in rows]
            return list(dict.fromkeys(db_admins + ENV_SUPER_ADMIN_IDS))
        
        elif role_prefix == 'admin_acc':
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'admin_acc' ORDER BY user_id")
            acc_admins = [row[0] for row in await cursor.fetchall()]
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'super_admin' ORDER BY user_id")
            super_admins = [row[0] for row in await cursor.fetchall()]
            # Сначала ответственные за аксессуары, затем супер-админы (без set(),
            # чтобы порядок был стабильным — важно для get_responsible_admin_for_category).
            return list(dict.fromkeys(acc_admins + super_admins + ENV_SUPER_ADMIN_IDS))
        
        elif role_prefix == 'admin_tech':
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'admin_tech' ORDER BY user_id")
            tech_admins = [row[0] for row in await cursor.fetchall()]
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'super_admin' ORDER BY user_id")
            super_admins = [row[0] for row in await cursor.fetchall()]
            return list(dict.fromkeys(tech_admins + super_admins + ENV_SUPER_ADMIN_IDS))
        
        elif role_prefix == 'admin_tradein':
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'admin_tradein' ORDER BY user_id")
            tradein_admins = [row[0] for row in await cursor.fetchall()]
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'super_admin' ORDER BY user_id")
            super_admins = [row[0] for row in await cursor.fetchall()]
            return list(dict.fromkeys(tradein_admins + super_admins + ENV_SUPER_ADMIN_IDS))
        
        elif role_prefix == 'admin_complaint':
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'admin_complaint' ORDER BY user_id")
            complaint_admins = [row[0] for row in await cursor.fetchall()]
            cursor = await db.execute("SELECT user_id FROM users WHERE role = 'super_admin' ORDER BY user_id")
            super_admins = [row[0] for row in await cursor.fetchall()]
            return list(dict.fromkeys(complaint_admins + super_admins + ENV_SUPER_ADMIN_IDS))
        
        else:
            cursor = await db.execute("SELECT user_id FROM users WHERE role = ?", (role_prefix,))
            rows = await cursor.fetchall()
            return [row[0] for row in rows]


async def get_category_role_admins_only(category: str) -> list[int]:
    """Только админы роли категории (без супер-админов) — для «ответственного»."""
    role_prefix = CLAIM_CATEGORY_ADMIN_ROLE.get(category)
    if not role_prefix:
        return []
    async with get_connection() as db:
        cursor = await db.execute(
            "SELECT user_id FROM users WHERE role = ? ORDER BY user_id",
            (role_prefix,)
        )
        return [row[0] for row in await cursor.fetchall()]


async def get_all_admins_list() -> dict:
    async with get_connection() as db:
        result = {
            'super_admin': [],
            'admin_tech': [],
            'admin_acc': [],
            'admin_tradein': [],
            'admin_complaint': []
        }
        
        cursor = await db.execute(
            "SELECT user_id, role FROM users WHERE role = 'super_admin' ORDER BY user_id"
        )
        for row in await cursor.fetchall():
            result['super_admin'].append((row[0], None))
        
        cursor = await db.execute(
            "SELECT user_id, role FROM users WHERE role = 'admin_tech' ORDER BY user_id"
        )
        for row in await cursor.fetchall():
            result['admin_tech'].append((row[0], None))
        
        cursor = await db.execute(
            "SELECT user_id, role FROM users WHERE role = 'admin_acc' ORDER BY user_id"
        )
        for row in await cursor.fetchall():
            result['admin_acc'].append((row[0], None))
        
        cursor = await db.execute(
            "SELECT user_id, role FROM users WHERE role = 'admin_tradein' ORDER BY user_id"
        )
        for row in await cursor.fetchall():
            result['admin_tradein'].append((row[0], None))
        
        cursor = await db.execute(
            "SELECT user_id, role FROM users WHERE role = 'admin_complaint' ORDER BY user_id"
        )
        for row in await cursor.fetchall():
            result['admin_complaint'].append((row[0], None))
        
        return result

async def get_stats_overview():
    async with get_connection() as db:
        cursor_total = await db.execute("SELECT COUNT(*) FROM claims")
        total = (await cursor_total.fetchone())[0]
        cursor_pending = await db.execute("SELECT COUNT(*) FROM claims WHERE status = 'pending'")
        pending = (await cursor_pending.fetchone())[0]
        cursor_resolved = await db.execute(
            "SELECT COUNT(*) FROM claims WHERE status IN ('approved', 'rejected', 'repair', 'quality_check')"
        )
        resolved = (await cursor_resolved.fetchone())[0]
        return {'total': total, 'pending': pending, 'resolved': resolved}

async def get_stats_by_points():
    # Единый агрегирующий запрос вместо 1 + 5*N отдельных COUNT(*) на точку продаж
    # (устраняет N+1 проблему при росте числа торговых точек и заявок).
    async with get_connection() as db:
        cursor = await db.execute("""
            SELECT
                user_id,
                SUM(CASE WHEN sub_category = 'ПТВ' THEN 1 ELSE 0 END) AS ptv,
                SUM(CASE WHEN sub_category = 'Новое устройство' THEN 1 ELSE 0 END) AS new_dev,
                SUM(CASE WHEN category = 'acc' THEN 1 ELSE 0 END) AS acc,
                SUM(CASE WHEN category = 'tradein' THEN 1 ELSE 0 END) AS tradein,
                SUM(CASE WHEN category = 'complaint' THEN 1 ELSE 0 END) AS complaint,
                COUNT(*) AS total
            FROM claims
            GROUP BY user_id
            HAVING total > 0
            ORDER BY total DESC
        """)
        rows = await cursor.fetchall()
        stats_list = [
            {
                'user_id': row[0],
                'name': f"ТТ #{row[0]}",
                'ptv': row[1],
                'new': row[2],
                'acc': row[3],
                'tradein': row[4],
                'complaint': row[5],
                'total': row[6],
            }
            for row in rows
        ]
        return stats_list

async def get_pending_claims():
    """Просроченные заявки (>2ч без решения).

    Возвращает также tg_name/client_name вместе с user_id одним запросом
    (без отдельного N+1 обращения к Telegram через bot.get_chat для каждой
    строки) — этого достаточно, чтобы вызывающий код построил кликабельную
    ссылку на автора заявки (см. handlers/super_admin.py: stats_pending).
    """
    async with get_connection() as db:
        cursor = await db.execute("""
            SELECT id, display_id, user_id, category, sub_category, created_at, tg_name, client_name
            FROM claims 
            WHERE status = 'pending' 
            AND (julianday('now') - julianday(created_at)) * 24 > 2 
            ORDER BY created_at ASC
        """)
        return await cursor.fetchall()

EXPORT_HEADERS = [
    "ID", "Номер заявки", "User ID", "Категория", "Подкатегория",
    "Название товара", "Дефект", "IMEI", "Дата покупки", "Пожелание клиента",
    "Статус", "Ответственный", "Клиент", "Имя в Telegram",
    "Способ оплаты", "Предложение конкурента",
    "Дата создания", "Дата решения", "Время решения",
]


async def _fetch_export_rows(days: int | None = None) -> list:
    """Готовые к выгрузке (Excel/CSV) строки отчёта — единая точка формирования,
    чтобы оба формата экспорта не расходились в наборе и порядке колонок.

    days=None — выгрузка за всё время, иначе — только заявки не старше N дней
    (используется период-фильтром на панели супер-админа).
    """
    from utils.export_format import (
        extract_imei,
        extract_product_name,
        format_category_ru,
        format_defect_for_export,
        format_status_ru,
        format_datetime_export,
        split_datetime_export,
    )

    base_query = """
        SELECT
            c.id, c.display_id, c.user_id, c.category, c.sub_category, c.brand,
            c.defect_desc, c.purchase_date, c.client_wish, c.status, c.admin_name,
            c.client_name, c.tg_name, c.payment_method, c.competitor_offer,
            c.created_at, rh.resolved_at
        FROM claims c
        LEFT JOIN (
            SELECT claim_id, MAX(changed_at) AS resolved_at
            FROM claim_history
            WHERE old_status = 'pending' AND new_status != 'pending'
            GROUP BY claim_id
        ) rh ON rh.claim_id = c.id
    """
    if days:
        query = base_query + " WHERE date(c.created_at) >= date('now', ?) ORDER BY c.created_at DESC"
        params = (f"-{int(days)} days",)
    else:
        query = base_query + " ORDER BY c.created_at DESC"
        params = ()

    async with get_connection() as db:
        cursor = await db.execute(query, params)
        rows = await cursor.fetchall()

    export_rows = []
    for row in rows:
        (
            claim_id,
            display_id,
            user_id,
            category,
            sub_category,
            brand,
            defect_desc,
            purchase_date,
            client_wish,
            status,
            admin_name,
            client_name,
            tg_name,
            payment_method,
            competitor_offer,
            created_at,
            resolved_at,
        ) = row

        # created_at / resolved_at в БД — UTC (CURRENT_TIMESTAMP); для отчёта
        # показываем Asia/Yekaterinburg (см. utils.tz / assume_utc=True).
        resolved_date, resolved_time = split_datetime_export(resolved_at, assume_utc=True)
        export_rows.append([
            claim_id,
            display_id,
            user_id,
            format_category_ru(category),
            sub_category,
            extract_product_name(brand, category),
            format_defect_for_export(defect_desc, category, brand),
            extract_imei(brand, defect_desc, category),
            purchase_date,
            client_wish,
            format_status_ru(status),
            admin_name,
            client_name,
            tg_name,
            payment_method or "",
            competitor_offer or "",
            format_datetime_export(created_at, assume_utc=True),
            resolved_date,
            resolved_time,
        ])
    return export_rows


async def export_stats_to_excel(days: int | None = None) -> bytes:
    if not OPENPYXL_AVAILABLE:
        return b"Error: openpyxl library not installed. Please run 'pip install openpyxl'"

    export_rows = await _fetch_export_rows(days)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по заявкам"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_bg = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    col_widths = [5, 10, 10, 14, 18, 25, 35, 18, 14, 20, 16, 18, 20, 22, 16, 20, 18, 14, 10]

    for col_num, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_bg
        cell.alignment = center_align
        cell.border = thin_border
        if col_num <= len(col_widths):
            ws.column_dimensions[chr(64 + col_num)].width = col_widths[col_num - 1]

    for row_idx, export_row in enumerate(export_rows, 2):
        for col_num, value in enumerate(export_row, 1):
            cell_value = "" if value is None else str(value)
            cell = ws.cell(row=row_idx, column=col_num, value=cell_value)
            cell.border = thin_border
            cell.alignment = left_align

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

async def export_stats_to_csv(days: int | None = None) -> bytes:
    export_rows = await _fetch_export_rows(days)

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
    writer.writerow(EXPORT_HEADERS)
    for row in export_rows:
        writer.writerow(["" if value is None else value for value in row])
    return output.getvalue().encode('utf-8-sig')

async def get_claims_count() -> int:
    async with get_connection() as db:
        cursor = await db.execute("SELECT COUNT(*) FROM claims")
        return (await cursor.fetchone())[0]

async def get_archive_count() -> int:
    async with get_connection() as db:
        cursor = await db.execute("SELECT COUNT(*) FROM claims_archive")
        return (await cursor.fetchone())[0]

async def clear_all_claims():
    async with get_connection() as db:
        await db.execute("BEGIN IMMEDIATE")
        try:
            await db.execute("DELETE FROM claims")
            await db.execute("DELETE FROM claims_archive")
            await db.execute("DELETE FROM claim_history")
            await db.execute("DELETE FROM chat_messages")
            await db.execute("UPDATE claim_counters SET last_number = 0 WHERE category = 'tech'")
            await db.execute("UPDATE claim_counters SET last_number = 0 WHERE category = 'acc'")
            await db.execute("UPDATE claim_counters SET last_number = 0 WHERE category = 'tradein'")
            await db.execute("UPDATE claim_counters SET last_number = 0 WHERE category = 'complaint'")
            await db.commit()
        except Exception:
            await db.rollback()
            raise


# ==========================================
# ТАЙМЕР ОТСУТСТВИЯ ОТВЕТА НА ЗАЯВКУ (5/10/15 МИН)
# ==========================================
# taken_at/taken_by — момент «реакции» на заявку для остановки таймера
# напоминаний 5/10/15 мин. Сейчас выставляется из чата заявки
# (utils/claim_timer_service.stop_claim_timer_if_needed ← handlers/chat.py).
# reminder_stage — до какой стадии напоминаний уже дошла заявка.

async def mark_claim_taken(claim_id: int, admin_id: int) -> bool:
    """Фиксирует реакцию на заявку (остановка таймера напоминаний), только если
    ещё не зафиксирована. Возвращает True, если именно этот вызов установил
    taken_at (SQL: `WHERE taken_at IS NULL`)."""
    async with get_connection() as db:
        cursor = await db.execute(
            "UPDATE claims SET taken_at = CURRENT_TIMESTAMP, taken_by = ? WHERE id = ? AND taken_at IS NULL",
            (admin_id, claim_id)
        )
        await db.commit()
        return cursor.rowcount > 0


async def get_overdue_claims_for_stage(stage: int, minutes: int) -> list[dict]:
    """Заявки без реакции (не взяты в работу), которым уже пора напомнить
    на стадии `stage` (1/2/3 — соответствуют порогам 5/10/15 минут).

    `reminder_stage = stage - 1` гарантирует, что стадия обрабатывается ровно
    один раз по очереди (стадия 2 не сработает, пока не отправлена стадия 1),
    поэтому claim_timer_loop должен обходить стадии по возрастанию (1, 2, 3).
    """
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            """
            SELECT * FROM claims
            WHERE status = 'pending'
              AND taken_at IS NULL
              AND reminder_stage = ?
              AND (julianday('now') - julianday(created_at)) * 24 * 60 >= ?
            ORDER BY created_at ASC
            """,
            (stage - 1, minutes)
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


async def set_claim_reminder_stage(claim_id: int, stage: int) -> None:
    async with get_connection() as db:
        await db.execute(
            "UPDATE claims SET reminder_stage = ? WHERE id = ?",
            (stage, claim_id)
        )
        await db.commit()


async def save_claim_admin_card(claim_id: int, chat_id: int, message_id: int) -> None:
    """Сохраняет message_id карточки заявки в чате админа для reply-напоминаний
    таймера (аксы / техника / Trade-in / корректировка остатков)."""
    if not claim_id or not chat_id or not message_id:
        return
    async with get_connection() as db:
        await db.execute(
            """
            INSERT INTO claim_admin_cards (claim_id, chat_id, message_id)
            VALUES (?, ?, ?)
            ON CONFLICT(claim_id, chat_id) DO UPDATE SET
                message_id = excluded.message_id,
                created_at = CURRENT_TIMESTAMP
            """,
            (int(claim_id), int(chat_id), int(message_id))
        )
        await db.commit()


async def get_claim_admin_card_message_id(claim_id: int, chat_id: int) -> int | None:
    """message_id исходной карточки заявки в чате chat_id, либо None."""
    async with get_connection() as db:
        cursor = await db.execute(
            "SELECT message_id FROM claim_admin_cards WHERE claim_id = ? AND chat_id = ?",
            (int(claim_id), int(chat_id))
        )
        row = await cursor.fetchone()
        return int(row[0]) if row else None


async def get_claim_admin_card_chat_ids(claim_id: int) -> list[int]:
    """Все chat_id, куда уже ушла карточка этой заявки (для напоминаний)."""
    async with get_connection() as db:
        cursor = await db.execute(
            "SELECT chat_id FROM claim_admin_cards WHERE claim_id = ? ORDER BY chat_id",
            (int(claim_id),)
        )
        return [int(row[0]) for row in await cursor.fetchall()]


async def get_stage1_reminder_recipients(category: str, claim_id: int) -> list[int]:
    """Получатели напоминания на 5-й минуте.

    1) Все админы роли категории (admin_acc / admin_tech / …).
    2) Если роли никого нет — все, кому уже отправили карточку заявки.
    3) Если карточек тоже нет — все супер-админы.
    """
    role_admins = await get_category_role_admins_only(category)
    if role_admins:
        return list(dict.fromkeys(role_admins))

    card_chats = await get_claim_admin_card_chat_ids(claim_id)
    if card_chats:
        return list(dict.fromkeys(card_chats))

    return list(dict.fromkeys(await get_admins_by_role('super_admin')))


async def get_responsible_admin_for_category(category: str) -> int | None:
    """Первый админ роли категории; если роли нет — первый супер-админ.

    Оставлен для совместимости; таймер 5 мин использует get_stage1_reminder_recipients.
    """
    role_admins = await get_category_role_admins_only(category)
    if role_admins:
        return role_admins[0]
    supers = await get_admins_by_role('super_admin')
    return supers[0] if supers else None


async def get_overdue_claims_detailed() -> list[dict]:
    """Заявки без реакции по НОВОЙ логике таймера (5/10/15 минут), в отличие
    от get_pending_claims() (просрочка > 2ч, используется в handlers/super_admin.py).

    Не изменяет и не заменяет get_pending_claims() — отдельная функция ДЛЯ
    БУДУЩЕГО подключения в разделе "⏳ Просроченные заявки" (см. отчёт агента),
    возвращает дополнительно taken_at/reminder_stage для более точного
    отображения текущей стадии просрочки каждой заявки.
    """
    async with get_connection() as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            """
            SELECT id, display_id, user_id, category, sub_category, created_at,
                   tg_name, client_name, taken_at, taken_by, reminder_stage,
                   (julianday('now') - julianday(created_at)) * 24 * 60 AS minutes_since_created
            FROM claims
            WHERE status = 'pending'
              AND taken_at IS NULL
              AND reminder_stage >= 1
            ORDER BY created_at ASC
            """
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]
